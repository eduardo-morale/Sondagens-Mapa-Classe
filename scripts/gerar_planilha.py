
import random, datetime, string, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# MODO TESTE: dados 100% ficticios, sem chamada ao Databricks
random.seed(42)

MUNICIPIOS = ["Sao Paulo","Campinas","Santos","Sorocaba","Ribeirao Preto",
              "Guarulhos","Osasco","Bauru","Marilia","Presidente Prudente"]
REDES  = ["Estadual","Municipal"]
SERIES = [1,2,3,4,5]
HIP    = ["Pre-silabica","Silabica sem valor","Silabica com valor","Silabica alfabetica","Alfabetica"]
MAT    = ["Nivel 1","Nivel 2","Nivel 3","Nivel 4"]
PROD   = ["Nivel 1","Nivel 2","Nivel 3"]
DES    = [(1,"DE Sao Paulo Centro"),(2,"DE Campinas Leste"),(3,"DE Santos"),(4,"DE Sorocaba")]

def _ra():     return str(random.randint(100_000_000, 999_999_999))
def _cie():    return str(random.randint(1_000_000, 9_999_999))
def _escola(): return f"EE Prof. {chr(random.randint(65,90))}.{''.join(random.choices(string.ascii_uppercase,k=5))}"
def _data(d=30): return (datetime.date.today()-datetime.timedelta(days=random.randint(0,d))).isoformat()

print("Gerando dados ficticios...")
resultados = {}

# 1. Evolucao Diaria
cols = ["data","rede","municipio","serie","nr_ra","cie","de_id","de_nome","nome_escola"]
rows = [[_data(),random.choice(REDES),random.choice(MUNICIPIOS),random.choice(SERIES),
         _ra(),_cie(),*(de:=random.choice(DES)),_escola()] for _ in range(200)]
resultados["Evolucao Diaria"] = (cols, rows)

# 2. Hipotese de Escrita
cols = ["rede","serie","serie_label","nivel","ordem_nivel","municipio","nr_ra","cie","de_id","de_nome","nome_escola"]
rows = [[random.choice(REDES),(s:=random.choice(SERIES[:2])),f"{s} Ano",(n:=random.choice(HIP)),
         HIP.index(n)+1,random.choice(MUNICIPIOS),_ra(),_cie(),*(de:=random.choice(DES)),_escola()]
        for _ in range(200)]
resultados["Hipotese de Escrita"] = (cols, rows)

# 3. Nivel de Matematica
cols = ["rede","serie","serie_label","nivel","ordem_nivel","municipio","nr_ra","cie","de_id","de_nome","nome_escola"]
rows = [[random.choice(REDES),(s:=random.choice(SERIES)),f"{s} Ano",(n:=random.choice(MAT)),
         MAT.index(n)+1,random.choice(MUNICIPIOS),_ra(),_cie(),*(de:=random.choice(DES)),_escola()]
        for _ in range(200)]
resultados["Nivel de Matematica"] = (cols, rows)

# 4. Producao Textual
cols = ["serie","serie_label","nivel","ordem_nivel","municipio","nr_ra","cie","de_id","de_nome","nome_escola"]
rows = [[(s:=random.choice([3,4,5])),f"{s} Ano",(n:=random.choice(PROD)),PROD.index(n)+1,
         random.choice(MUNICIPIOS),_ra(),_cie(),*(de:=random.choice(DES)),_escola()]
        for _ in range(150)]
resultados["Producao Textual"] = (cols, rows)

# 5. Cobertura Sondagens
cols = ["rede","municipio","serie","nr_ra","cie","de_id","datafinalizado",
        "hipotese_escrita_resultado","nivel_matematica_resultado","producao_textual_resultado",
        "nome_escola","de_nome"]
rows = [[random.choice(REDES),random.choice(MUNICIPIOS),random.choice(SERIES),
         _ra(),_cie(),*(de:=random.choice(DES)),_data(),
         random.choice(HIP),random.choice(MAT),random.choice(PROD),_escola(),de[1]]
        for _ in range(200)]
resultados["Cobertura Sondagens"] = (cols, rows)

# 6. Cobertura Esperada
cols = ["rede","serie","cie","nome_escola","municipio","de_id","de_nome",
        "alunos_esperados","alunos_sondados","status_escola"]
rows = []
for _ in range(80):
    de=random.choice(DES); esp=random.randint(20,40); sond=random.randint(0,esp+5)
    st="Nao iniciou" if sond==0 else ("Concluida" if sond>=esp*0.95 else "Parcial")
    rows.append([random.choice(REDES),random.choice(SERIES),_cie(),_escola(),
                 random.choice(MUNICIPIOS),de[0],de[1],esp,sond,st])
resultados["Cobertura Esperada"] = (cols, rows)

# 7. Pct Conclusao
cols = ["categoria","ordem","pct_conclusao"]
rows = [["Total",1,round(random.uniform(40,75),1)],
        ["Estadual",2,round(random.uniform(45,80),1)],
        ["Municipal",3,round(random.uniform(30,65),1)]]
resultados["Pct Conclusao"] = (cols, rows)

# 8. Sem Sondagem
cols = ["rede","cie","nome_escola","municipio","de_id","de_nome",
        "alunos_esperados","alunos_sondados","status_escola"]
rows = [[random.choice(REDES),_cie(),_escola(),random.choice(MUNICIPIOS),
         *(de:=random.choice(DES)),random.randint(20,40),0,"Nao iniciou"] for _ in range(30)]
resultados["Sem Sondagem"] = (cols, rows)

# 9. Inconsistencias
cols = ["rede","serie","serie_label","cie","nome_escola","municipio","de_id","de_nome",
        "alunos_esperados","alunos_sondados","excedente","pct_sondagem"]
rows = []
for _ in range(20):
    de=random.choice(DES); s=random.choice(SERIES); esp=random.randint(20,35)
    sond=esp+random.randint(1,10)
    rows.append([random.choice(REDES),s,f"{s} Ano",_cie(),_escola(),random.choice(MUNICIPIOS),
                 de[0],de[1],esp,sond,sond-esp,round(sond*100/esp,1)])
resultados["Inconsistencias"] = (cols, rows)

print(f"  {sum(len(r) for _,r in resultados.values())} linhas em {len(resultados)} tabelas")

# Gerar Excel (identico a versao de producao)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)

wb = Workbook()
wb.remove(wb.active)

for nome, (cols, rows) in resultados.items():
    ws = wb.create_sheet(title=nome[:31])
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = "A2"
    for ci, col in enumerate(cols, 1):
        max_len = max(len(col), max((len(str(r[ci-1])) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

ts = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
wb.create_sheet(title="_meta").cell(row=1, column=1, value=f"[TESTE] Gerado em: {ts} (UTC)")

os.makedirs("output", exist_ok=True)
wb.save("output/mapa_classe_sondagens.xlsx")
print(f"Excel salvo (MODO TESTE) — {sum(len(r) for _,r in resultados.values())} linhas totais")
